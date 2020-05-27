import openpyxl as xl
import utils

wb = xl.load_workbook("Template for hours worked.xlsx")
ws = wb["Sheet1"]
total_hours = 0

days = int(input("How many days did you work this 2 week period?: "))
for row in range(1,days+1):
    ws.insert_rows(row + 5)
    cellHours = ws.cell(5+row, 2)
    cellDate = ws.cell(5+row, 1)
    cellDay = ws.cell(5+row, 3)
    cellHours.value = int(input(f"Hours worked on day {row}: "))
    total_hours = total_hours + cellHours.value
    cellDate.value = input(f"Enter Date of day {row} in DD/MM/YYYY: ")
    cellDay.value = utils.day_from_date(cellDate.value)


cell_hours_total = ws.cell(ws.max_row,2)
cell_hours_total.value = total_hours
cellEnd = ws.cell(3,3)
payroll_end = input("Payroll period end DD/MM/YYYY: ")
date_frag = utils.number_to_words_date(payroll_end)
cellEnd.value = payroll_end

doc_name = f"{date_frag[1]} {date_frag[2]} {date_frag[3]} PAYROLL.xlsx"
wb.save(doc_name)